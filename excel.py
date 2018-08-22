import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from period import Period


class Excel:
    """
    Raporty z Trawersa w formie arkuszy Excela
    """

    def __init__(self):
        self.__period = Period()

    def sale(self, sales):
        """Dane sprzedaży bieżącego miesiąca"""
        daily_sales, monthly_sales = sales
        filename = 'sprzedaż - ' + self.__period.today_date() + '.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'sprzedaż'

        # wartości dzienne sprzedaży
        ws['A1'] = 'Sprzedaż - ' + self.__period.month_year_str()
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        ws.cell(row=2, column=1)
        ws['A3'] = 'data'
        ws['B3'] = 'wino'
        ws['C3'] = 'cydry'
        ws['D3'] = 'wódki'
        ws['E3'] = 'soki'
        ws['F3'] = 'netto'
        ws['G3'] = 'brutto'
        ws['A3'].alignment = Alignment(horizontal='center')
        for c in range(2, 8):
            ws.cell(row=3, column=c).alignment = Alignment(horizontal='right')
        for c in range(1, 8):
            ws.cell(row=3, column=c).font = Font(bold=True)
            ws.cell(row=3, column=c).border = Border(bottom=Side(border_style='thin'))

        for row in daily_sales:
            line = (row[0], row[1], row[3], row[5], row[7], row[9], row[10])
            ws.append(line)
        days = len(daily_sales)
        for r in range(days):
            ws.cell(row=r + 4, column=1).number_format = 'dd-mm-yyyy'
            ws.cell(row=r + 4, column=1).alignment = Alignment(horizontal='center')
            ws.cell(row=r + 4, column=2).number_format = '#,##0'
            ws.cell(row=r + 4, column=3).number_format = '#,##0'
            ws.cell(row=r + 4, column=4).number_format = '#,##0'
            ws.cell(row=r + 4, column=5).number_format = '#,##0'
            ws.cell(row=r + 4, column=6).number_format = '#,##0.00'
            ws.cell(row=r + 4, column=7).number_format = '#,##0.00'
        for c in range(7):
            ws.cell(row=days + 3, column=c + 1).border = Border(bottom=Side(border_style='thin'))
        for c in range(7):
            ws.column_dimensions[get_column_letter(c + 1)].width = 12

        # dane sumaryczne
        ws.cell(row=days + 5, column=1, value='razem:')
        ws.cell(row=days + 5, column=2, value='wino')
        ws.cell(row=days + 5, column=3, value='cydry')
        ws.cell(row=days + 5, column=4, value='wódki')
        ws.cell(row=days + 5, column=5, value='soki')
        ws.cell(row=days + 5, column=6, value='netto')
        ws.cell(row=days + 5, column=7, value='brutto')
        ws.cell(row=days + 5, column=1).alignment = Alignment(horizontal='center')
        for c in range(2, 8):
            ws.cell(row=days + 5, column=c).alignment = Alignment(horizontal='right')
        for c in range(1, 8):
            ws.cell(row=days + 5, column=c).font = Font(bold=True)
            ws.cell(row=days + 5, column=c).border = Border(bottom=Side(border_style='thin'))

        wine_quantity = int(monthly_sales[0])
        wine_value = float(monthly_sales[1])
        wine_price = 0
        if wine_quantity != 0:
            wine_price = wine_value / wine_quantity
        cider_quantity = int(monthly_sales[2])
        cider_value = float(monthly_sales[3])
        cider_price = 0
        if cider_quantity != 0:
            cider_price = cider_value / cider_quantity
        vodka_quantity = int(monthly_sales[4])
        vodka_value = float(monthly_sales[5])
        vodka_price = 0
        if vodka_quantity != 0:
            vodka_price = vodka_value / vodka_quantity
        juice_quantity = int(monthly_sales[6])
        juice_value = float(monthly_sales[7])
        juice_price = 0
        if juice_quantity != 0:
            juice_price = juice_value / juice_quantity
        net_value = float(monthly_sales[8])
        gross_value = float(monthly_sales[9])

        ws.cell(row=days + 6, column=1, value='- ilość')
        ws.cell(row=days + 7, column=1, value='- wartość')
        ws.cell(row=days + 8, column=1, value='- śred. cena')
        ws.cell(row=days + 6, column=2, value=wine_quantity).number_format = '#,##0'
        ws.cell(row=days + 7, column=2, value=wine_value).number_format = '#,##0.00'
        ws.cell(row=days + 8, column=2, value=wine_price).number_format = '#,##0.00'
        ws.cell(row=days + 6, column=3, value=cider_quantity).number_format = '#,##0'
        ws.cell(row=days + 7, column=3, value=cider_value).number_format = '#,##0.00'
        ws.cell(row=days + 8, column=3, value=cider_price).number_format = '#,##0.00'
        ws.cell(row=days + 6, column=4, value=vodka_quantity).number_format = '#,##0'
        ws.cell(row=days + 7, column=4, value=vodka_value).number_format = '#,##0.00'
        ws.cell(row=days + 8, column=4, value=vodka_price).number_format = '#,##0.00'
        ws.cell(row=days + 6, column=5, value=juice_quantity).number_format = '#,##0'
        ws.cell(row=days + 7, column=5, value=juice_value).number_format = '#,##0.00'
        ws.cell(row=days + 8, column=5, value=juice_price).number_format = '#,##0.00'
        ws.cell(row=days + 6, column=6, value=net_value).number_format = '#,##0.00'
        ws.cell(row=days + 6, column=6, ).alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=days + 6, start_column=6, end_row=days + 8, end_column=6)
        ws.cell(row=days + 6, column=7, value=gross_value).number_format = '#,##0.00'
        ws.cell(row=days + 6, column=7, ).alignment = Alignment(vertical='center')
        ws.merge_cells(start_row=days + 6, start_column=7, end_row=days + 8, end_column=7)

        wb.save(filename)
        return filename
